# xlsx_to_two_jsons_gui.py
# Drag-and-drop an .xlsx and export:
#   <prefix>_structure.json  (constants + formula text)
#   <prefix>_values.json     (cached values for formula cells only)
#
# Dependencies:
#   pip install tkinterdnd2
#
# Notes:
# - Cached values come from what Excel last saved. If many are null, open in Excel, recalc, save, re-export.

import json
import re
import threading
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

import tkinter as tk
from tkinter import filedialog, messagebox

from tkinterdnd2 import DND_FILES, TkinterDnD


# ---------- core export logic ----------

def safe_json(v):
    if v is None:
        return None
    if isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, datetime):
        return v.isoformat()
    try:
        return str(v)
    except Exception:
        return repr(v)


def col_letters_to_index(col: str) -> int:
    col = col.upper().strip()
    n = 0
    for ch in col:
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"Bad column letters: {col}")
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def parse_a1_range(a1: str):
    a1 = a1.strip()
    m = re.fullmatch(r"([A-Za-z]+)(\d+):([A-Za-z]+)(\d+)", a1)
    if not m:
        raise ValueError(f'Range must look like "A1:DN500". Got: {a1}')
    c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
    min_col = col_letters_to_index(c1)
    max_col = col_letters_to_index(c2)
    min_row = r1
    max_row = r2
    if min_row > max_row or min_col > max_col:
        raise ValueError(f"Range is inverted: {a1}")
    return min_row, min_col, max_row, max_col


def extract_defined_names(wb):
    out = []
    try:
        items = wb.defined_names.items()
    except Exception:
        return out

    for name, defn in items:
        defs = defn if isinstance(defn, (list, tuple)) else [defn]
        for d in defs:
            out.append({
                "name": name,
                "localSheetId": getattr(d, "localSheetId", None),
                "refers_to": getattr(d, "attr_text", None) or safe_json(d),
                "comment": getattr(d, "comment", None),
            })
    return out


def build_structure_json(xlsx_path: Path, a1_range: str):
    min_row, min_col, max_row, max_col = parse_a1_range(a1_range)

    wb = load_workbook(filename=str(xlsx_path), data_only=False, keep_links=True)
    defined_names = extract_defined_names(wb)

    formula_addr_map = {}
    sheets = []
    total_cells_kept = 0
    total_formula_cells = 0

    for ws in wb.worksheets:
        formula_addrs = set()
        cells = []

        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                if cell.data_type == "f":
                    formula_addrs.add(cell.coordinate)
                    cells.append({
                        "addr": cell.coordinate,
                        "value": None,
                        "formula": str(cell.value),  # str() handles ArrayFormula safely
                        "number_format": safe_json(cell.number_format),
                    })
                else:
                    if cell.value is None:
                        continue
                    cells.append({
                        "addr": cell.coordinate,
                        "value": safe_json(cell.value),
                        "formula": None,
                        "number_format": safe_json(cell.number_format),
                    })

        formula_addr_map[ws.title] = formula_addrs
        total_cells_kept += len(cells)
        total_formula_cells += len(formula_addrs)

        sheets.append({
            "name": ws.title,
            "state": ws.sheet_state,
            "dimensions": a1_range,
            "cell_count": len(cells),
            "formula_cell_count": len(formula_addrs),
            "cells": cells,
        })

    export = {
        "source_file": str(xlsx_path),
        "range_exported": a1_range,
        "sheet_count": len(wb.worksheets),
        "defined_name_count": len(defined_names),
        "defined_names": defined_names,
        "total_cells_kept": total_cells_kept,
        "total_formula_cells": total_formula_cells,
        "sheets": sheets,
        "notes": {
            "structure": "Non-empty constants + formula text for formula cells.",
            "calc": "No recalculation performed.",
        }
    }
    return export, formula_addr_map


def build_values_json(xlsx_path: Path, a1_range: str, formula_addr_map):
    min_row, min_col, max_row, max_col = parse_a1_range(a1_range)

    wb = load_workbook(filename=str(xlsx_path), data_only=True, keep_links=True)

    sheets = []
    total_formula_values = 0
    total_missing_cached_values = 0

    for ws in wb.worksheets:
        formula_addrs = formula_addr_map.get(ws.title, set())
        cells = []
        missing = 0

        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                if cell.coordinate not in formula_addrs:
                    continue
                v = safe_json(cell.value)
                if v is None:
                    missing += 1
                cells.append({
                    "addr": cell.coordinate,
                    "value": v,
                    "number_format": safe_json(cell.number_format),
                })

        total_formula_values += len(cells)
        total_missing_cached_values += missing

        sheets.append({
            "name": ws.title,
            "state": ws.sheet_state,
            "dimensions": a1_range,
            "formula_value_count": len(cells),
            "missing_cached_value_count": missing,
            "cells": cells,
        })

    export = {
        "source_file": str(xlsx_path),
        "range_exported": a1_range,
        "sheet_count": len(wb.worksheets),
        "total_formula_values": total_formula_values,
        "total_missing_cached_values": total_missing_cached_values,
        "sheets": sheets,
        "notes": {
            "values": "Cached values for formula cells only (what Excel last saved).",
        }
    }
    return export


def export_two_jsons(xlsx_path: Path, a1_range: str, outdir: Path, prefix: str):
    structure_export, formula_addr_map = build_structure_json(xlsx_path, a1_range)
    values_export = build_values_json(xlsx_path, a1_range, formula_addr_map)

    outdir.mkdir(parents=True, exist_ok=True)
    structure_path = outdir / f"{prefix}_structure.json"
    values_path = outdir / f"{prefix}_values.json"

    with structure_path.open("w", encoding="utf-8") as f:
        json.dump(structure_export, f, indent=2, ensure_ascii=False)

    with values_path.open("w", encoding="utf-8") as f:
        json.dump(values_export, f, indent=2, ensure_ascii=False)

    return structure_path, values_path, structure_export, values_export


# ---------- GUI ----------

def parse_dnd_files(data: str):
    # Windows drag-drop sometimes wraps paths with spaces in {...}
    # Example: '{C:\Path With Spaces\file.xlsx}'
    paths = []
    token = ""
    in_brace = False
    for ch in data:
        if ch == "{":
            in_brace = True
            token = ""
        elif ch == "}":
            in_brace = False
            if token:
                paths.append(token)
                token = ""
        elif ch == " " and not in_brace:
            if token:
                paths.append(token)
                token = ""
        else:
            token += ch
    if token:
        paths.append(token)
    return [p.strip() for p in paths if p.strip()]


class App(TkinterDnD.Tk):
    def __init__(self):
        super().__init__()
        self.title("XLSX → Two JSONs (Structure + Values)")
        self.geometry("860x520")

        self.xlsx_path = tk.StringVar(value="")
        self.range_str = tk.StringVar(value="A1:DN500")
        self.outdir_str = tk.StringVar(value="")
        self.prefix_str = tk.StringVar(value="")

        self._build_ui()

    def _build_ui(self):
        top = tk.Frame(self)
        top.pack(fill="x", padx=12, pady=10)

        # Drop zone
        drop = tk.Label(
            top,
            text="Drag & drop an .xlsx here\n(or click Browse)",
            relief="ridge",
            borderwidth=3,
            padx=20,
            pady=18,
            anchor="center",
        )
        drop.pack(fill="x")

        drop.drop_target_register(DND_FILES)
        drop.dnd_bind("<<Drop>>", self.on_drop)

        btns = tk.Frame(top)
        btns.pack(fill="x", pady=8)

        tk.Button(btns, text="Browse…", command=self.browse).pack(side="left")
        tk.Button(btns, text="Set Output Folder…", command=self.set_outdir).pack(side="left", padx=8)

        form = tk.Frame(self)
        form.pack(fill="x", padx=12)

        def row(label, var, width=80):
            r = tk.Frame(form)
            r.pack(fill="x", pady=3)
            tk.Label(r, text=label, width=16, anchor="w").pack(side="left")
            e = tk.Entry(r, textvariable=var, width=width)
            e.pack(side="left", fill="x", expand=True)
            return e

        row("XLSX Path", self.xlsx_path)
        row("Range", self.range_str)
        row("Output Dir", self.outdir_str)
        row("Prefix", self.prefix_str)

        actions = tk.Frame(self)
        actions.pack(fill="x", padx=12, pady=10)

        self.run_btn = tk.Button(actions, text="Run Export", command=self.run_export)
        self.run_btn.pack(side="left")

        tk.Button(actions, text="Clear Log", command=self.clear_log).pack(side="left", padx=8)

        self.status = tk.Label(actions, text="", anchor="w")
        self.status.pack(side="left", padx=12)

        # Log box
        self.log = tk.Text(self, height=18, wrap="none")
        self.log.pack(fill="both", expand=True, padx=12, pady=8)

        self._log("Ready. Drop an .xlsx or click Browse.")

    def _log(self, msg: str):
        self.log.insert("end", msg + "\n")
        self.log.see("end")

    def clear_log(self):
        self.log.delete("1.0", "end")

    def on_drop(self, event):
        paths = parse_dnd_files(event.data)
        if not paths:
            return
        p = Path(paths[0])
        if p.suffix.lower() != ".xlsx":
            messagebox.showerror("Wrong file", "Drop an .xlsx file.")
            return
        self.set_file(p)

    def browse(self):
        p = filedialog.askopenfilename(
            title="Select .xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
        )
        if not p:
            return
        self.set_file(Path(p))

    def set_outdir(self):
        d = filedialog.askdirectory(title="Choose output folder")
        if not d:
            return
        self.outdir_str.set(d)

    def set_file(self, p: Path):
        self.xlsx_path.set(str(p))
        if not self.outdir_str.get():
            self.outdir_str.set(str(p.parent))
        if not self.prefix_str.get():
            self.prefix_str.set(p.stem)
        self._log(f"Selected: {p}")

    def run_export(self):
        xlsx = self.xlsx_path.get().strip()
        if not xlsx:
            messagebox.showerror("Missing file", "Pick or drop an .xlsx first.")
            return

        xlsx_path = Path(xlsx)
        if not xlsx_path.exists():
            messagebox.showerror("Not found", f"File not found:\n{xlsx_path}")
            return

        a1_range = self.range_str.get().strip()
        outdir = Path(self.outdir_str.get().strip() or xlsx_path.parent)
        prefix = self.prefix_str.get().strip() or xlsx_path.stem

        self.run_btn.config(state="disabled")
        self.status.config(text="Running…")
        self._log(f"Running export on: {xlsx_path}")
        self._log(f"Range: {a1_range}")
        self._log(f"Output: {outdir} (prefix: {prefix})")

        def worker():
            try:
                structure_path, values_path, s_export, v_export = export_two_jsons(
                    xlsx_path=xlsx_path,
                    a1_range=a1_range,
                    outdir=outdir,
                    prefix=prefix,
                )
                self.after(0, lambda: self._on_success(structure_path, values_path, s_export, v_export))
            except Exception as e:
                self.after(0, lambda: self._on_error(e))

        threading.Thread(target=worker, daemon=True).start()

    def _on_success(self, structure_path, values_path, s_export, v_export):
        self._log("DONE ✅")
        self._log(f"Structure JSON: {structure_path}")
        self._log(f"Values JSON:    {values_path}")
        self._log(f"Sheets: {s_export.get('sheet_count')}")
        self._log(f"Defined names: {s_export.get('defined_name_count')}")
        self._log(f"Formula cells: {s_export.get('total_formula_cells')}")
        self._log(f"Formula cached values missing (null): {v_export.get('total_missing_cached_values')}")
        self.status.config(text="Done")
        self.run_btn.config(state="normal")

    def _on_error(self, e: Exception):
        self._log("ERROR ❌")
        self._log(str(e))
        self.status.config(text="Error")
        self.run_btn.config(state="normal")


if __name__ == "__main__":
    App().mainloop()
